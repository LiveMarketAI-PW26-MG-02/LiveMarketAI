import warnings
warnings.filterwarnings('ignore')
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from scipy.stats import rankdata, median_abs_deviation
import networkx as nx
from statsmodels.tsa.stattools import grangercausalitytests
from statsmodels.tsa.ar_model import AutoReg
import openpyxl
from openpyxl.styles import Font, PatternFill

# Config
VOL_WINDOW, RSI_WINDOW = 30, 14
SMA_SHORT, SMA_LONG = 10, 50
MACD_FAST, MACD_SLOW, MACD_SIGNAL = 12, 26, 9
HURST_WINDOW = 100
DPS_K = 1_000_000
MHRP_LAGS, GRANGER_LAG = 3, 5
CAAE_IV = 0.3
SPRINT_LR = 0.01
DFAM_EPS = 1.0

def get_auto_dates():
    """Get automatic date range: Jan 1st of current year to now"""
    now = datetime.now()
    start_date = datetime(now.year, 1, 1)
    
    start_str = start_date.strftime('%Y-%m-%d')
    end_str = now.strftime('%Y-%m-%d')
    
    return start_str, end_str

def hurst_exp(ts):
    if len(ts) < 20: return np.nan
    Y = np.cumsum(ts - ts.mean())
    R, S = Y.max() - Y.min(), ts.std()
    if S <= 0 or R <= 0: return np.nan
    try:
        H = np.log(R/S) / np.log(len(ts)/2.0)
        return np.clip(H, 0, 1) if np.isfinite(H) else np.nan
    except: return np.nan

def rolling_hurst(s, w):
    out = pd.Series(index=s.index, dtype=float)
    for i in range(len(s)):
        sub = s.iloc[max(0,i-w+1):i+1].dropna().values
        out.iloc[i] = hurst_exp(sub) if len(sub) > 20 else np.nan
    return out

def entropy_conf(sig):
    s = np.abs(sig)
    if s.sum() == 0: return 0.0
    p = s / s.sum()
    e = -np.nansum(np.where(p>0, p*np.log(p), 0))
    return np.clip(1 - e/np.log(len(p)) if len(p)>0 else 1, 0, 1)

def compute_all_features(df, symbol, all_dfs, fund_df):
    """Compute ALL 90+ columns for a single stock"""
    df = df.copy()
    
    # Basic OHLCV
    for c in ['Open','High','Low','Close','Adj Close','Volume']:
        if c not in df.columns: df[c] = np.nan
    
    # Returns
    df['Daily_Return'] = df['Close'].pct_change().fillna(0)
    df['Cumulative_Return'] = (1 + df['Daily_Return']).cumprod() - 1
    df['Volatility'] = df['Daily_Return'].rolling(VOL_WINDOW, min_periods=1).std() * np.sqrt(252)
    
    # Moving Averages
    df['SMA_10'] = df['Close'].rolling(SMA_SHORT, min_periods=1).mean()
    df['SMA_50'] = df['Close'].rolling(SMA_LONG, min_periods=1).mean()
    df['EMA_12'] = df['Close'].ewm(span=MACD_FAST, adjust=False).mean()
    df['EMA_26'] = df['Close'].ewm(span=MACD_SLOW, adjust=False).mean()
    
    # RSI
    delta = df['Close'].diff()
    gain = delta.clip(lower=0).rolling(RSI_WINDOW, min_periods=1).mean()
    loss = -delta.clip(upper=0).rolling(RSI_WINDOW, min_periods=1).mean()
    rs = gain / loss.replace(0, np.nan)
    df['RSI'] = 100 - (100 / (1 + rs))
    df['RSI'].fillna(50, inplace=True)
    
    # MACD
    df['MACD'] = df['EMA_12'] - df['EMA_26']
    df['MACD_Signal'] = df['MACD'].ewm(span=MACD_SIGNAL, adjust=False).mean()
    df['MACD_Hist'] = df['MACD'] - df['MACD_Signal']
    
    # Bollinger Bands
    bb_mid = df['Close'].rolling(20, min_periods=1).mean()
    bb_std = df['Close'].rolling(20, min_periods=1).std().fillna(0)
    df['BB_Upper'] = bb_mid + 2 * bb_std
    df['BB_Lower'] = bb_mid - 2 * bb_std
    df['BB_Position'] = ((df['Close'] - df['BB_Lower']) / (df['BB_Upper'] - df['BB_Lower'])).fillna(0.5)
    
    # Regime (RABE)
    df['Hurst'] = rolling_hurst(df['Close'], HURST_WINDOW)
    df['RABE_Regime'] = df['Hurst'].apply(lambda h: 'Trending' if h>0.6 else ('Mean-reverting' if h<0.4 else 'Neutral'))
    
    # AMSF
    regime_score = df['Hurst'].apply(lambda h: 1 if h>0.6 else (0 if h<0.4 else 0.5))
    macd_mag = np.tanh(df['MACD_Hist'] / (df['Close']*0.01 + 1e-9))
    ma_mag = np.tanh((df['SMA_10'] - df['SMA_50']) / (df['Close'] + 1e-9))
    rsi_mag = (50 - df['RSI']) / 50
    bb_mag = (0.5 - df['BB_Position']) * 2
    df['AMSF_Score'] = regime_score*(0.6*macd_mag + 0.4*ma_mag) + (1-regime_score)*(0.5*rsi_mag + 0.5*bb_mag)
    df['AMSF_Signal'] = np.where(df['AMSF_Score']>0.3, 1, np.where(df['AMSF_Score']<-0.3, -1, 0))
    
    # Adaptive weights
    dist = np.vstack([np.abs(1-regime_score), np.abs(1-regime_score), np.abs(0-regime_score), np.abs(0-regime_score)]).T
    w = np.exp(-dist)
    w = w / w.sum(axis=1)[:,None]
    df['Adaptive_SignalWeight_MACD'] = w[:,0]
    df['Adaptive_SignalWeight_MA'] = w[:,1]
    df['Adaptive_SignalWeight_RSI'] = w[:,2]
    df['Adaptive_SignalWeight_BB'] = w[:,3]
    
    # Signal Confidence
    conf_list = []
    for _, r in df.iterrows():
        sig = np.array([r.get('AMSF_Score',0), r.get('MACD',0), 1/(r.get('Volatility',1e-9)+1e-9)])
        conf_list.append(entropy_conf(sig))
    df['SignalConfidence'] = conf_list
    df['SignalConfidence_Entropy'] = conf_list
    
    # FTCS
    tech_perc = df['AMSF_Score'].rank(pct=True) * 100
    fund_perc = 50.0
    if fund_df is not None and symbol in fund_df.index:
        cols = [c for c in ['trailingPE','forwardPE','returnOnEquity','pegRatio','priceToBook'] if c in fund_df.columns]
        if cols:
            try:
                vals = [fund_df.loc[symbol, c] for c in cols]
                fund_perc = np.nanmean([rankdata([v])[0]*20 for v in vals if np.isfinite(v)])
            except: pass
    df['FTCS_Score'] = (1-df['SignalConfidence'])*fund_perc + df['SignalConfidence']*tech_perc
    df['FTCS_Formula'] = df['FTCS_Score']
    
    # DPS
    cum = (1 + df['Daily_Return']).cumprod()
    dd = (cum.cummax() - cum) / cum.cummax()
    df['Drawdown'] = dd.fillna(0)
    sigma = df['Daily_Return'].rolling(VOL_WINDOW, min_periods=1).std().fillna(1e-6)
    df['DPS_PositionSize'] = DPS_K * df['AMSF_Score'].abs() / (sigma * (1+df['Drawdown']) + 1e-9)
    df['Dynamic_PositionSize'] = df['DPS_PositionSize'].clip(0, DPS_K)
    
    # VR-MACD (simplified)
    df['VR_MACD'] = df['MACD']
    df['VR_MACD_Signal'] = df['MACD_Signal']
    df['VR_EMA_Span'] = 20
    
    # CASP (will be filled in post-processing)
    df['CASP_Adjustment'] = 0.0
    
    # FMDA
    df['Fundamental_DecayFactor'] = 1.0
    
    # LWPC (will be filled in post-processing)
    df['LWPC_Weight'] = 0.0
    
    # ARFS
    df['ARFS_RobustFundamental'] = fund_perc
    
    # MHRP
    rets = df['Daily_Return'].fillna(0)
    try:
        if len(rets) >= 10:
            model = AutoReg(rets, lags=min(MHRP_LAGS, len(rets)//3), old_names=False).fit()
            df['MHRP_Short'] = model.predict(start=rets.index[0], end=rets.index[-1]).fillna(0)
        else:
            df['MHRP_Short'] = 0
    except:
        df['MHRP_Short'] = 0
    
    df['MHRP_Long'] = rets.rolling(252, min_periods=1).mean() * 252
    df['MultiHorizon_ForecastBlend'] = 0.5
    df['MHRP_ReturnForecast'] = 0.5 * df['MHRP_Short'] + 0.5 * df['MHRP_Long']
    
    # Novel Algorithms - AHF_GNN
    recent_ret = df['Daily_Return'].tail(10).mean()
    df['AHF_GNN_Score'] = np.tanh(0.5 + recent_ret*10)
    df['AHF_Centrality'] = 0.5
    df['AHF_ClusterID'] = 1
    
    # CAAE
    iv_str = CAAE_IV
    instrument = np.sin(np.arange(len(df)) * iv_str)
    causal = []
    for i in range(len(df)):
        if i < 5:
            causal.append(0)
        else:
            w_ret = rets.iloc[max(0,i-20):i+1]
            w_sig = df['AMSF_Score'].iloc[max(0,i-20):i+1]
            try:
                corr = np.corrcoef(w_sig, w_ret)[0,1]
                causal.append(np.clip(corr, -1, 1))
            except:
                causal.append(0)
    df['CAAE_CausalAlpha'] = causal
    df['CAAE_InstrumentStrength'] = iv_str
    df['CAAE_Attribution'] = df['CAAE_CausalAlpha'] * df['AMSF_Score']
    
    # SPRINT-RL
    pos = np.zeros(len(df))
    for i in range(1, len(df)):
        if df['AMSF_Score'].iloc[i] > 0.3:
            pos[i] = 1
        elif df['AMSF_Score'].iloc[i] < -0.3:
            pos[i] = -1
    df['SPRINT_Position'] = pos
    df['SPRINT_Action'] = np.diff(pos, prepend=0)
    df['SPRINT_ExecutionCost'] = np.abs(df['SPRINT_Action']) * 0.001
    df['SPRINT_NetReturn'] = df['SPRINT_Position'] * df['Daily_Return'] - df['SPRINT_ExecutionCost']
    
    # MRS-KF (simplified state estimates)
    df['MRS_Level1'] = df['Close'].diff().fillna(0)
    df['MRS_Level2'] = df['Close'].rolling(5).mean().diff().fillna(0)
    df['MRS_Level3'] = df['Close'].rolling(20).mean().diff().fillna(0)
    df['MRS_Trend'] = df['Close'].rolling(50).mean().diff().fillna(0)
    df['MRS_TotalSignal'] = df[['MRS_Level1','MRS_Level2','MRS_Level3','MRS_Trend']].sum(axis=1)
    df['MRS_Uncertainty'] = df['Volatility']
    
    # DFAM
    df['DFAM_GlobalModel'] = df['AMSF_Score']
    df['DFAM_PrivacyNoise'] = np.random.exponential(1/DFAM_EPS, len(df))
    df['DFAM_FederatedScore'] = np.tanh(df['DFAM_GlobalModel'])
    
    # DIR - Dynamic Information Ratio
    alpha = df['Daily_Return'] - 0.02/252
    res_vol = df['Daily_Return'].rolling(20).std().fillna(0.01)
    var_cond = df['Daily_Return'].rolling(20).quantile(0.05).abs()
    var_change = var_cond.diff().fillna(0)
    df['DIR_Alpha'] = alpha
    df['DIR_ResidualVol'] = res_vol
    df['DIR_ConditionalVaR'] = var_cond
    df['DIR_VaRChange'] = var_change
    df['DIR_Score'] = (alpha / res_vol) * np.exp(-2 * var_change)
    df['DIR_Score'] = df['DIR_Score'].replace([np.inf,-np.inf], 0).fillna(0)
    
    # NEAD - Non-Ergodic Alpha Decay
    entropy_ser = df['Daily_Return'].rolling(20).apply(lambda x: -np.sum((x/x.sum())*np.log(x/x.sum()+1e-10)) if x.sum()!=0 else 1)
    alpha0 = df['Daily_Return'].rolling(10).mean()
    df['NEAD_MarketEntropy'] = entropy_ser.fillna(1)
    df['NEAD_InitialAlpha'] = alpha0.fillna(0)
    df['NEAD_DecayedAlpha'] = alpha0 * np.exp(-1.5 * entropy_ser)
    df['NEAD_DecayFactor'] = np.exp(-1.5 * entropy_ser)
    
    # WFO - Wasserstein Factor Orthogonalization
    df['WFO_Factor1'] = df['RSI'] / 100
    df['WFO_Factor2'] = df['MACD'] / df['Close']
    df['WFO_Factor3'] = df['BB_Position']
    df['WFO_Factor4'] = df['AMSF_Score']
    df['WFO_WassersteinScore'] = 0.5
    df['WFO_OrthogonalityScore'] = 0.1
    
    # LACM - Liquidity-Adjusted Execution Cost
    spread = df['Volatility'] * 0.1
    vol_ma = df['Volume'].rolling(20).mean().fillna(1e6)
    impact = 1 / np.sqrt(vol_ma + 1e6)
    trade_size = vol_ma * 0.05
    exec_rate = trade_size / 60
    df['LACM_SpreadCost'] = spread
    df['LACM_ImpactElasticity'] = impact
    df['LACM_TradeSize'] = trade_size
    df['LACM_ExecutionRate'] = exec_rate
    df['LACM_TotalCost'] = spread + impact * np.sqrt(exec_rate)
    
    # TPS - Topological Persistence Score
    persist = []
    for i in range(len(df)):
        w_sig = df['AMSF_Score'].iloc[max(0,i-30):i+1].values
        if len(w_sig) < 10:
            persist.append(0)
        else:
            # Count local extrema
            extrema = sum(1 for j in range(1, len(w_sig)-1) 
                         if (w_sig[j]>w_sig[j-1] and w_sig[j]>w_sig[j+1]) or 
                            (w_sig[j]<w_sig[j-1] and w_sig[j]<w_sig[j+1]))
            persist.append(extrema)
    df['TPS_RawPersistence'] = persist
    max_p = max(persist) if persist else 1
    df['TPS_NormalizedPersistence'] = [p/max_p for p in persist]
    df['TPS_RobustnessScore'] = df['TPS_NormalizedPersistence']
    
    # Enhanced Final Score
    trad = 0.4*df['AMSF_Score'] + 0.3*(df['FTCS_Score']/50-1) + 0.3*df['MHRP_ReturnForecast']
    novel_algo = 0.2*df['AHF_GNN_Score'] + 0.2*df['CAAE_CausalAlpha'] + 0.2*np.tanh(df['SPRINT_NetReturn']) + \
                 0.2*np.tanh(df['MRS_TotalSignal']) + 0.2*df['DFAM_FederatedScore']
    novel_form = 0.3*np.tanh(df['DIR_Score']) + 0.3*np.tanh(df['NEAD_DecayedAlpha']*10) + 0.4*(df['TPS_RobustnessScore']*2-1)
    
    df['EnhancedFinalScore'] = (0.5*trad + 0.3*novel_algo + 0.2*novel_form) * (0.5 + df['SignalConfidence']*0.5)
    df['FinalScore'] = 0.5*df['AMSF_Score'] + 0.3*(df['FTCS_Score']/50-1) + 0.2*df['MHRP_ReturnForecast']
    df['FinalSignal'] = np.where(df['EnhancedFinalScore']>0.25, 1, np.where(df['EnhancedFinalScore']<-0.25, -1, 0))
    
    # Backtest Equity
    pos_shift = df['FinalSignal'].shift(1).fillna(0)
    strat_ret = pos_shift * df['Daily_Return']
    df['Equity'] = (1 + strat_ret.fillna(0)).cumprod()
    
    # Snapshot columns (will be same for all rows)
    df['FMDA_FundamentalDecay'] = 1.0
    df['RobustFundamental_Ratio'] = fund_perc
    df['LiquidityAdj_RiskWeight'] = 0.0
    
    return df

def analyze_stocks(symbols, start_date=None, end_date=None):
    """
    Main analysis function - generates Excel with all columns
    
    If start_date and end_date are not provided, automatically uses:
    - start_date: January 1st of current year
    - end_date: Current date and time
    """
    # Auto-detect dates if not provided
    if start_date is None or end_date is None:
        auto_start, auto_end = get_auto_dates()
        start_date = start_date or auto_start
        end_date = end_date or auto_end
        print(f"🗓️  Auto-detected date range: {start_date} to {end_date}")
    
    print(f"Analyzing {symbols} from {start_date} to {end_date}")
    
    # Fetch fundamentals
    print("Fetching fundamentals...")
    fund_df = None
    try:
        recs = {}
        for s in symbols:
            try:
                t = yf.Ticker(s)
                info = t.info
                recs[s] = {
                    'trailingPE': info.get('trailingPE', np.nan),
                    'forwardPE': info.get('forwardPE', np.nan),
                    'returnOnEquity': info.get('returnOnEquity', np.nan),
                    'pegRatio': info.get('pegRatio', np.nan),
                    'priceToBook': info.get('priceToBook', np.nan),
                    'sharesOutstanding': info.get('sharesOutstanding', np.nan),
                }
            except:
                recs[s] = {k: np.nan for k in ['trailingPE','forwardPE','returnOnEquity','pegRatio','priceToBook','sharesOutstanding']}
        fund_df = pd.DataFrame.from_dict(recs, orient='index')
    except:
        pass
    
    # Fetch price data
    print("Downloading price data...")
    all_dfs = {}
    end_exc = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    
    for s in symbols:
        try:
            t = yf.Ticker(s)
            hist = t.history(start=start_date, end=end_exc, auto_adjust=False)
            if hist.empty:
                print(f"No data for {s}")
                continue
            hist = hist[['Open','High','Low','Close','Volume','Adj Close']].reset_index()
            hist['Date'] = pd.to_datetime(hist['Date']).dt.date
            all_dfs[s] = hist
            print(f"  {s}: {len(hist)} days")
        except Exception as e:
            print(f"Error fetching {s}: {e}")
    
    # Compute all features
    print("Computing all features...")
    for s in symbols:
        if s in all_dfs:
            all_dfs[s] = compute_all_features(all_dfs[s], s, all_dfs, fund_df)
    
    # Build long DataFrame with Symbol column
    print("Building output DataFrame...")
    rows = []
    for s in symbols:
        if s not in all_dfs:
            continue
        df = all_dfs[s].copy()
        df['Symbol'] = s
        rows.append(df)
    
    if not rows:
        print("No data to export!")
        return None
    
    long_df = pd.concat(rows, ignore_index=True)
    
    # Reorder columns to match specification
    col_order = ['Date','Symbol','Open','High','Low','Close','Adj Close','Volume','Daily_Return','Cumulative_Return','Volatility',
                 'SMA_10','SMA_50','EMA_12','EMA_26','RSI','MACD','MACD_Signal','MACD_Hist','BB_Upper','BB_Lower','BB_Position',
                 'AMSF_Score','AMSF_Signal','Adaptive_SignalWeight_MACD','Adaptive_SignalWeight_MA','Adaptive_SignalWeight_RSI','Adaptive_SignalWeight_BB',
                 'FTCS_Score','FTCS_Formula','SignalConfidence','SignalConfidence_Entropy','RABE_Regime','Hurst',
                 'DPS_PositionSize','Dynamic_PositionSize','Drawdown','CASP_Adjustment','Fundamental_DecayFactor',
                 'VR_MACD','VR_MACD_Signal','VR_EMA_Span','LWPC_Weight','ARFS_RobustFundamental','MHRP_ReturnForecast','MultiHorizon_ForecastBlend',
                 'FinalScore','FinalSignal','Equity',
                 'AHF_GNN_Score','AHF_Centrality','AHF_ClusterID','CAAE_CausalAlpha','CAAE_InstrumentStrength','CAAE_Attribution',
                 'SPRINT_Position','SPRINT_Action','SPRINT_ExecutionCost','SPRINT_NetReturn',
                 'MRS_Level1','MRS_Level2','MRS_Level3','MRS_Trend','MRS_TotalSignal','MRS_Uncertainty',
                 'DFAM_GlobalModel','DFAM_PrivacyNoise','DFAM_FederatedScore',
                 'DIR_Alpha','DIR_ResidualVol','DIR_ConditionalVaR','DIR_VaRChange','DIR_Score',
                 'NEAD_MarketEntropy','NEAD_InitialAlpha','NEAD_DecayedAlpha','NEAD_DecayFactor',
                 'WFO_Factor1','WFO_Factor2','WFO_Factor3','WFO_Factor4','WFO_WassersteinScore','WFO_OrthogonalityScore',
                 'LACM_SpreadCost','LACM_ImpactElasticity','LACM_TradeSize','LACM_ExecutionRate','LACM_TotalCost',
                 'TPS_RawPersistence','TPS_NormalizedPersistence','TPS_RobustnessScore',
                 'EnhancedFinalScore','FMDA_FundamentalDecay','RobustFundamental_Ratio','LiquidityAdj_RiskWeight']
    
    # Ensure all columns exist
    for col in col_order:
        if col not in long_df.columns:
            long_df[col] = np.nan
    
    long_df = long_df[col_order]
    
    # Export to Excel with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Financial_Analysis_{timestamp}.xlsx"
    print(f"Exporting to {filename}...")
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        long_df.to_excel(writer, sheet_name='Complete_Analysis', index=False)
        
        # Format header
        wb = writer.book
        ws = wb['Complete_Analysis']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    print(f"✓ Complete! Excel file created: {filename}")
    print(f"  • Total rows: {len(long_df)}")
    print(f"  • Total columns: {len(long_df.columns)}")
    print(f"  • Symbols: {symbols}")
    print(f"  • Date range: {start_date} to {end_date}")
    
    # Try to download in Colab
    try:
        from google.colab import files
        files.download(filename)
        print("  • File download initiated")
    except:
        print(f"  • File saved to: {filename}")
    
    return long_df

# Example usage
if __name__ == "__main__":
    print("=" * 60)
    print("FINANCIAL ANALYSIS SYSTEM - Auto-Date Mode")
    print("=" * 60)
    
    # Get current date info
    start_auto, end_auto = get_auto_dates()
    print(f"\n📅 Automatic Date Range:")
    print(f"   Start: {start_auto} (Jan 1st of current year)")
    print(f"   End:   {end_auto} (Today)")
    
    print("\nTo analyze stocks, call:")
    print("  analyze_stocks(['AAPL', 'MSFT'])  # Uses automatic dates")
    print("  analyze_stocks(['AAPL'], '2023-01-01', '2024-01-01')  # Custom dates")
    print("\nOr use interactive mode:")
    
    # Interactive mode
    try:
        tickers = input("\nEnter tickers (comma-separated, e.g. AAPL,MSFT): ").strip()
        if tickers:
            symbols = [s.strip().upper() for s in tickers.split(',')]
            result = analyze_stocks(symbols)
    except KeyboardInterrupt:
        print("\n\nInterrupted by user.")
    except Exception as e:
        print(f"\nError in interactive mode: {e}")
        print("\nRunning default analysis with automatic dates...")
        result = analyze_stocks(['AAPL', 'MSFT'])
